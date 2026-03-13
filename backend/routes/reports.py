from __future__ import annotations

from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any

from flask import Blueprint, jsonify, request, send_file
from sqlalchemy import func

from database import db
from models import Trip, Vehicle
from routes.auth import require_role

reports_bp = Blueprint('reports', __name__)

PLAN_WEEK = 105000
LOW_DAY_LT = 14500
NORM_DAY_MIN = 14500
NORM_DAY_MAX = 16000


def _parse_iso_date(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except Exception:
        return None


def _dow_ru_short(d: date) -> str:
    # Monday=0..Sunday=6
    names = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']
    return names[d.weekday()]


def _day_label(d: date) -> str:
    return f"{d.strftime('%d.%m')} ({_dow_ru_short(d)})"


def _clip_range(start: date, end: date) -> tuple[date, date, bool]:
    if end < start:
        start, end = end, start
    used_end = min(end, start + timedelta(days=6))
    return start, used_end, used_end != end


def _classify_week_total(total: float) -> str:
    if abs(total - PLAN_WEEK) < 1e-9:
        return 'Норма (план)'
    if total > PLAN_WEEK:
        return 'Высокий доход'
    return 'Низкий доход'


def _count_buckets(day_amounts: list[float]) -> tuple[int, int, int]:
    low = sum(1 for v in day_amounts if v < LOW_DAY_LT)
    norm = sum(1 for v in day_amounts if NORM_DAY_MIN <= v <= NORM_DAY_MAX)
    high = sum(1 for v in day_amounts if v > NORM_DAY_MAX)
    return low, norm, high


def _build_preview(start: date, end: date, plate: str | None) -> dict[str, Any]:
    start, used_end, truncated = _clip_range(start, end)
    days = [start + timedelta(days=i) for i in range((used_end - start).days + 1)]
    day_keys = [d.isoformat() for d in days]

    q = (
        db.session.query(
            Vehicle.license_plate.label('plate'),
            Trip.trip_date.label('trip_date'),
            func.coalesce(func.sum(Trip.price_with_vat), 0.0).label('income'),
        )
        .join(Vehicle, Trip.vehicle_id == Vehicle.id)
        .filter(Trip.trip_date >= start, Trip.trip_date <= used_end)
    )

    if plate:
        q = q.filter(Vehicle.license_plate == plate)

    q = q.group_by(Vehicle.license_plate, Trip.trip_date)

    rows_map: dict[str, dict[str, float]] = {}
    for plate_val, trip_date_val, income in q.all():
        if not plate_val or not trip_date_val:
            continue
        plate_str = str(plate_val)
        trip_day = trip_date_val.isoformat() if isinstance(trip_date_val, date) else str(trip_date_val)
        rows_map.setdefault(plate_str, {})[trip_day] = float(income or 0.0)

    totals_by_day = [0.0 for _ in days]

    rows: list[dict[str, Any]] = []
    for plate_str in sorted(rows_map.keys()):
        per_day = [float(rows_map[plate_str].get(k, 0.0)) for k in day_keys]
        week_total = float(sum(per_day))
        deviation = week_total - PLAN_WEEK
        completion_pct = (week_total / PLAN_WEEK) if PLAN_WEEK else 0.0
        group = _classify_week_total(week_total)
        low_days, norm_days, high_days = _count_buckets(per_day)

        for i, v in enumerate(per_day):
            totals_by_day[i] += v

        rows.append(
            {
                'plate': plate_str,
                'dayAmounts': per_day,
                'weekTotal': week_total,
                'plan': PLAN_WEEK,
                'deviation': deviation,
                'completionPct': completion_pct,
                'group': group,
                'lowDays': low_days,
                'normDays': norm_days,
                'highDays': high_days,
            }
        )

    return {
        'days': [{'date': d.isoformat(), 'label': _day_label(d)} for d in days],
        'rows': rows,
        'totalsByDay': totals_by_day,
        'truncated': truncated,
        'used_end': used_end.isoformat(),
    }


@reports_bp.route('/reports/weekly-preview', methods=['GET'])
@require_role('admin', 'dispatcher', 'viewer')
def weekly_preview():
    start = _parse_iso_date(request.args.get('start'))
    end = _parse_iso_date(request.args.get('end'))
    plate = request.args.get('plate')

    if not start or not end:
        return jsonify({'error': 'start and end are required (YYYY-MM-DD)'}), 400

    data = _build_preview(start, end, plate)
    return jsonify(data), 200


@reports_bp.route('/reports/weekly-excel', methods=['GET'])
@require_role('admin', 'dispatcher', 'viewer')
def weekly_excel():
    start = _parse_iso_date(request.args.get('start'))
    end = _parse_iso_date(request.args.get('end'))
    plate = request.args.get('plate')

    if not start or not end:
        return jsonify({'error': 'start and end are required (YYYY-MM-DD)'}), 400

    preview = _build_preview(start, end, plate)

    try:
        from openpyxl import load_workbook
    except Exception as e:
        return jsonify({'error': f'openpyxl is required: {e}'}), 500

    template_path = Path(__file__).resolve().parents[2] / 'database temp' / 'otchet.xlsx'
    wb = load_workbook(template_path)
    ws = wb.worksheets[0]

    days = [datetime.strptime(d['date'], '%Y-%m-%d').date() for d in preview['days']]

    # Headers B1..H1 (7 columns)
    for i in range(7):
        col = 2 + i  # B=2
        if i < len(days):
            ws.cell(row=1, column=col).value = _day_label(days[i])
        else:
            ws.cell(row=1, column=col).value = ''

    # Clear old data area (safe: only A..H and J for rows 2..19).
    # Do NOT touch I/K/L/M/N/O/P because they contain formulas.
    for r in range(2, 20):
        for c in range(1, 9):
            ws.cell(row=r, column=c).value = None
        ws.cell(row=r, column=10).value = None

    # Fill rows
    excel_row = 2
    for row in preview['rows']:
        if excel_row > 19:
            break
        ws.cell(row=excel_row, column=1).value = row['plate']
        for i in range(7):
            v = row['dayAmounts'][i] if i < len(row['dayAmounts']) else 0
            ws.cell(row=excel_row, column=2 + i).value = float(v)
        ws.cell(row=excel_row, column=10).value = PLAN_WEEK  # J
        excel_row += 1

    # Totals row B20..H20
    totals = preview['totalsByDay']
    for i in range(7):
        ws.cell(row=20, column=2 + i).value = float(totals[i] if i < len(totals) else 0.0)

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    start_s = preview['days'][0]['date'] if preview['days'] else start.isoformat()
    end_s = preview['used_end']
    filename = f"otchet_{start_s}_{end_s}.xlsx"

    return send_file(
        out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )
