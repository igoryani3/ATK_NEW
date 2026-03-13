#!/usr/bin/env python3
"""
Import trips from Google Sheets to database
Each sheet represents one day (sheet name = day number)
Usage: python import_trips_from_sheets.py [--dry-run]
"""

import sys
import os
import argparse
from datetime import datetime, date
from io import BytesIO
import urllib.request
from openpyxl import load_workbook
from app import app
from database import db
from models import Trip, Contract, Customer, Region, TripType, Executor, Driver, Vehicle

# Configuration
SPREADSHEET_ID = os.environ.get('GOOGLE_SHEETS_SPREADSHEET_ID') or "1sD67A8G28eysHjh6pqTpsiHOanXnvSqWuJ2Ll0bUU2w"
MONTH_YEAR = (3, 2026)  # March 2026

# Column mapping (0-indexed)
COLUMN_INDICES = {
    'date': 0,              # Дата (ignore, use sheet name)
    'region': 1,            # Регион
    'contract': 2,          # Договор
    'customer': 3,          # Заказчик
    'name': 4,              # Название Маршрута
    'trip_type': 5,         # Тип заказа
    'passengers': 6,        # Количество мест
    'time_of_day': 7,       # Время суток
    'dispatch_time': 8,     # Подача
    'departure_time': 9,    # Выезд
    'route_start': 10,      # Начало маршрута
    'route_end': 11,        # Окончание маршрута
    'movement_type': 12,    # Тип рейса
    'executor': 13,         # Исполнитель
    'vehicle': 14,          # Транспорт
    'driver': 15,           # Водитель
    'driver_phone': 16,     # Номер водителя
    'price_no_vat': 17,     # Цена без НДС
    'price_with_vat': 18,   # Цена с НДС
    'weight': 19,           # Вес (ignore)
}


def fetch_xlsx_data():
    """Download spreadsheet as XLSX"""
    url = f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/export?format=xlsx"
    print(f"Downloading spreadsheet as XLSX...")

    try:
        with urllib.request.urlopen(url) as response:
            xlsx_content = response.read()
        return xlsx_content
    except Exception as e:
        print(f"Error downloading: {e}")
        sys.exit(1)


def parse_sheet_name_to_day(sheet_name):
    """Parse sheet name to get day number (e.g., '1' -> 1, '05' -> 5)"""
    try:
        # Try to extract number from sheet name
        name = sheet_name.strip()
        # Handle formats like "1", "01", "5", "05/03", "1 марта"
        if '/' in name:
            day_part = name.split('/')[0]
        else:
            day_part = ''.join(filter(str.isdigit, name))

        if day_part:
            return int(day_part)
    except:
        pass
    return None


def parse_float(value):
    """Parse float from cell value"""
    if value is None or value == '':
        return None
    try:
        if isinstance(value, (int, float)):
            return float(value)
        cleaned = str(value).replace(',', '.').replace(' ', '')
        return float(cleaned)
    except:
        return None


def parse_time(value, trip_date):
    """Parse time and combine with date"""
    if value is None or value == '':
        return None
    try:
        if isinstance(value, datetime):
            return datetime.combine(trip_date, value.time())

        time_str = str(value).strip()
        hours, minutes = map(int, time_str.split(':'))
        return datetime.combine(trip_date, datetime.min.time().replace(hour=hours, minute=minutes))
    except:
        return None


def get_cell_value(row, index):
    """Safely get cell value from row"""
    if index < len(row):
        val = row[index].value
        return str(val).strip() if val is not None else None
    return None


def find_contract(name):
    if not name:
        return None
    return Contract.query.filter_by(name=name).first()


def find_customer(name):
    if not name:
        return None
    return Customer.query.filter_by(name=name).first()


def find_region(name):
    if not name:
        return None
    return Region.query.filter_by(name=name).first()


def find_trip_type(name):
    if not name:
        return None
    return TripType.query.filter_by(name=name).first()


def find_executor(name):
    if not name:
        return None
    return Executor.query.filter_by(name=name).first()


def find_driver(name, executor_id):
    if not name:
        return None
    query = Driver.query.filter_by(full_name=name)
    if executor_id:
        query = query.filter_by(executor_id=executor_id)
    return query.first()


def find_vehicle(plate, executor_id):
    if not plate:
        return None
    query = Vehicle.query.filter_by(license_plate=plate)
    if executor_id:
        query = query.filter_by(executor_id=executor_id)
    return query.first()


def import_trips(dry_run=False):
    """Import trips from all sheets"""
    xlsx_content = fetch_xlsx_data()
    workbook = load_workbook(BytesIO(xlsx_content), data_only=True)

    month, year = MONTH_YEAR

    print(f"\nFound {len(workbook.sheetnames)} sheets: {workbook.sheetnames}")

    with app.app_context():
        stats = {
            'total': 0,
            'created': 0,
            'skipped': 0,
            'errors': 0,
            'sheets_processed': 0
        }

        for sheet_name in workbook.sheetnames:
            day = parse_sheet_name_to_day(sheet_name)

            if day is None or day < 1 or day > 31:
                print(f"\nSkipping sheet '{sheet_name}' - cannot parse day number")
                continue

            try:
                trip_date = date(year, month, day)
            except ValueError:
                print(f"\nSkipping sheet '{sheet_name}' - invalid date {year}-{month}-{day}")
                continue

            sheet = workbook[sheet_name]
            print(f"\n{'='*60}")
            print(f"Processing sheet '{sheet_name}' -> {trip_date.strftime('%Y-%m-%d')}")
            print(f"{'='*60}")

            # Skip header row, process data rows
            rows = list(sheet.iter_rows(min_row=2))
            sheet_trips = 0

            for row_idx, row in enumerate(rows, start=2):
                stats['total'] += 1

                try:
                    # Extract data
                    name = get_cell_value(row, COLUMN_INDICES['name'])
                    route_start = get_cell_value(row, COLUMN_INDICES['route_start'])
                    route_end = get_cell_value(row, COLUMN_INDICES['route_end'])

                    # Skip empty rows
                    if not route_start and not route_end:
                        stats['skipped'] += 1
                        continue

                    # Lookup references
                    contract_name = get_cell_value(row, COLUMN_INDICES['contract'])
                    customer_name = get_cell_value(row, COLUMN_INDICES['customer'])
                    region_name = get_cell_value(row, COLUMN_INDICES['region'])
                    trip_type_name = get_cell_value(row, COLUMN_INDICES['trip_type'])
                    executor_name = get_cell_value(row, COLUMN_INDICES['executor'])
                    driver_name = get_cell_value(row, COLUMN_INDICES['driver'])
                    vehicle_plate = get_cell_value(row, COLUMN_INDICES['vehicle'])

                    contract = find_contract(contract_name)
                    customer = find_customer(customer_name)
                    region = find_region(region_name)
                    trip_type = find_trip_type(trip_type_name)
                    executor = find_executor(executor_name)
                    driver = find_driver(driver_name, executor.id if executor else None)
                    vehicle = find_vehicle(vehicle_plate, executor.id if executor else None)

                    # Parse other fields
                    passengers_str = get_cell_value(row, COLUMN_INDICES['passengers'])
                    passengers_count = int(passengers_str) if passengers_str and passengers_str.isdigit() else None

                    time_of_day = get_cell_value(row, COLUMN_INDICES['time_of_day'])
                    movement_type = get_cell_value(row, COLUMN_INDICES['movement_type'])
                    driver_phone = get_cell_value(row, COLUMN_INDICES['driver_phone'])

                    dispatch_time_val = row[COLUMN_INDICES['dispatch_time']].value if COLUMN_INDICES['dispatch_time'] < len(row) else None
                    departure_time_val = row[COLUMN_INDICES['departure_time']].value if COLUMN_INDICES['departure_time'] < len(row) else None

                    dispatch_time = parse_time(dispatch_time_val, trip_date)
                    departure_time = parse_time(departure_time_val, trip_date)

                    price_no_vat_val = row[COLUMN_INDICES['price_no_vat']].value if COLUMN_INDICES['price_no_vat'] < len(row) else None
                    price_with_vat_val = row[COLUMN_INDICES['price_with_vat']].value if COLUMN_INDICES['price_with_vat'] < len(row) else None

                    price_no_vat = parse_float(price_no_vat_val)
                    price_with_vat = parse_float(price_with_vat_val)

                    # Create Trip
                    trip = Trip(
                        trip_date=trip_date,
                        name=name,
                        contract_id=contract.id if contract else None,
                        customer_id=customer.id if customer else None,
                        customer_name=customer_name if not customer else None,
                        region_id=region.id if region else None,
                        trip_type_id=trip_type.id if trip_type else None,
                        executor_id=executor.id if executor else None,
                        driver_id=driver.id if driver else None,
                        driver_name=driver_name if not driver else None,
                        driver_phone=driver_phone,
                        vehicle_id=vehicle.id if vehicle else None,
                        movement_type=movement_type,
                        passengers_count=passengers_count,
                        time_of_day=time_of_day,
                        dispatch_time=dispatch_time,
                        departure_time=departure_time,
                        route_start=route_start,
                        route_end=route_end,
                        price_no_vat=price_no_vat,
                        price_with_vat=price_with_vat
                    )

                    if not dry_run:
                        db.session.add(trip)

                    stats['created'] += 1
                    sheet_trips += 1

                    if stats['created'] % 100 == 0:
                        print(f"  Processed {stats['created']} trips total...")
                        if not dry_run:
                            db.session.commit()

                except Exception as e:
                    print(f"  Row {row_idx}: Error - {e}")
                    stats['errors'] += 1
                    continue

            print(f"  Imported {sheet_trips} trips from this sheet")
            stats['sheets_processed'] += 1

            if not dry_run:
                db.session.commit()

        # Final commit
        if not dry_run:
            db.session.commit()

        # Print statistics
        print("\n" + "="*60)
        print("IMPORT SUMMARY:")
        print(f"  Sheets processed: {stats['sheets_processed']}")
        print(f"  Total rows: {stats['total']}")
        print(f"  Created: {stats['created']}")
        print(f"  Skipped: {stats['skipped']}")
        print(f"  Errors: {stats['errors']}")
        print("="*60)

        if dry_run:
            print("\nDRY RUN - No data written to database")
        else:
            print(f"\nSuccessfully imported {stats['created']} trips!")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Import trips from Google Sheets')
    parser.add_argument('--dry-run', action='store_true', help='Run without writing to database')
    args = parser.parse_args()

    import_trips(dry_run=args.dry_run)
